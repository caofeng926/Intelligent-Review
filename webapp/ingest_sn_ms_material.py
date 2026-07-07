"""Shaanxi Medical Service Pricing 2021 -- Special Material Library ingest.

Source xlsx sheet '材料库' (sheet 8), 215 rows:
  - 40 top-level categories (HC001..HC040, 5-char) carrying a composite
    description of the consumables that fall under them.
  - 175 leaf items (HC00101..HC04003, 7-char) each naming a single
    consumable.  Parent link is via the 5-char prefix.

Publisher leaves bid_code blank; institutions are expected to fill it
at procurement time.
"""
from __future__ import annotations

import argparse
import os
import re
import sqlite3
import sys

try:
    import openpyxl
except ImportError:
    print("openpyxl is required: pip install openpyxl", file=sys.stderr)
    raise

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "kp.db")
XLSX = r"D:\Workspace\医保智审规则库\《陕西省医疗服务项目价格（2021版）》.xlsx"
SHEET_NAME = "材料库"

RE_CATEGORY = re.compile(r"^HC\d{3}$")  # top-level: HC001..HC040
RE_LEAF = re.compile(r"^HC\d{5}$")       # leaf:      HC00101..HC04003

C_FIN = 0
C_CODE = 1
C_NAME = 2
C_BID = 3


def _to_str(v):
    if v is None:
        return ""
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        return str(v)
    return str(v).strip()


def _level_for(code: str) -> int:
    if RE_CATEGORY.match(code):
        return 1
    if RE_LEAF.match(code):
        return 2
    return 0


def ingest(xlsx_path: str, conn: sqlite3.Connection, truncate: bool = False) -> int:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    cur = conn.cursor()
    if truncate:
        cur.execute("DELETE FROM sn_ms_material_codes")

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"sheet {SHEET_NAME!r} not found in xlsx")

    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))

    inserted = 0
    for row in rows:
        if not row or len(row) < 3:
            continue
        fin_raw = row[C_FIN] if C_FIN < len(row) else None
        code_raw = row[C_CODE] if C_CODE < len(row) else None
        name_raw = row[C_NAME] if C_NAME < len(row) else None
        bid_raw = row[C_BID] if C_BID < len(row) else None

        fin_class = _to_str(fin_raw)
        code = _to_str(code_raw)
        name = _to_str(name_raw)
        bid_code = _to_str(bid_raw)

        if not code:
            continue
        level = _level_for(code)
        if level == 0:
            continue
        if not name:
            continue

        if level == 2:
            p_code = code[:5]
        else:
            p_code = None

        level_path = code if level == 1 else f"{p_code}/{code}"

        cur.execute(
            """INSERT OR REPLACE INTO sn_ms_material_codes
               (code, p_code, level, fin_class, name, bid_code,
                sheet_name, level_path)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (code, p_code, level, fin_class or None, name,
             bid_code or None, SHEET_NAME, level_path),
        )
        inserted += 1

    conn.commit()
    wb.close()
    return inserted


def main(argv=None):
    p = argparse.ArgumentParser()
    p.add_argument("--xlsx", default=XLSX)
    p.add_argument("--db", default=None)
    p.add_argument("--truncate", action="store_true",
                   help="clear sn_ms_material_codes before insert")
    args = p.parse_args(argv)

    db_path = args.db or DB_PATH
    if not os.path.exists(args.xlsx):
        print(f"xlsx not found: {args.xlsx}", file=sys.stderr)
        return 2
    if not os.path.exists(db_path):
        print(f"db not found: {db_path}", file=sys.stderr)
        return 2

    sql_path = os.path.join(
        os.path.dirname(os.path.abspath(db_path)),
        "sn_ms_material_schema.sql",
    )
    conn = sqlite3.connect(db_path)
    if os.path.exists(sql_path):
        with open(sql_path, "r", encoding="utf-8") as fh:
            conn.executescript(fh.read())

    n = ingest(args.xlsx, conn, truncate=args.truncate)
    conn.close()
    print(f"inserted/updated {n} rows into sn_ms_material_codes")
    return 0


if __name__ == "__main__":
    sys.exit(main())