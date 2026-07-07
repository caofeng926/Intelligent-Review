"""Shaanxi Medical Service Pricing 2021 edition xlsx ingest.

Encoding level inference:
  L2 -> 2, L4 -> 3, L6 -> 4, L9 +/-alpha -> 5
"""
from __future__ import annotations

import argparse
import os
import re
import sqlite3
import sys

try:
    import openpyxl
except ImportError:
    print("openpyxl is required: pip install openpyxl", file=sys.stderr)
    raise

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "kp.db")
XLSX = r"D:\Workspace\医保智审规则库\《陕西省医疗服务项目价格（2021版）》.xlsx"

DIGITS_RE = re.compile(r"(\d+)")

C_FIN = 0
C_CODE = 1
C_NAME = 2
C_UNIT = 3
C_PRICE3 = 4
C_PRICE2 = 5
C_PRICE1 = 6
C_CONTENT = 7
C_EXCLUDE = 8
C_REMARK = 9


def _to_str(v):
    if v is None:
        return ""
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        return str(v)
    return str(v).strip()


def _norm_price(v):
    if v is None or v == "":
        return None
    s = _to_str(v)
    return s if s else None


def _infer(code: str):
    """Return (level, parent_code_digits_or_None, digits)."""
    if not code:
        return 0, None, None
    m = DIGITS_RE.match(code)
    if not m:
        return 0, None, None
    d = m.group(1)
    L = len(d)
    if L == 2:
        return 2, None, d
    if L == 4:
        return 3, d[:2], d
    if L == 6:
        return 4, d[:4], d
    if L >= 9:
        return 5, d[:4], d
    return 0, None, None


def ingest(xlsx_path: str, conn: sqlite3.Connection, truncate: bool = False) -> int:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    cur = conn.cursor()
    if truncate:
        cur.execute("DELETE FROM sn_ms_codes")

    inserted = 0
    sheet_names = [s for s in wb.sheetnames if s != "材料库"]
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        # First non-empty row of column A holds sheet title (e.g. "一、综合医疗服务类")
        sheet_title = sheet_name
        for r in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            if r and r[0]:
                v = _to_str(r[0])
                if v:
                    sheet_title = v
                    break

        # Ancestors cache: code -> name  (for level_path building)
        ancestors_name = {}

        rows = list(ws.iter_rows(values_only=True))
        n_rows = len(rows)
        # Skip header rows: row 0 (title), row 1 (note), row 2 (column headers), row 3 (sub-headers)
        # Data starts at row 4 (1-indexed). In 0-indexed: start at 3.
        for idx in range(3, n_rows):
            row = rows[idx]
            if not row or len(row) < 3:
                continue
            code_raw = row[C_CODE] if C_CODE < len(row) else None
            name_raw = row[C_NAME] if C_NAME < len(row) else None
            code = _to_str(code_raw)
            name = _to_str(name_raw)
            if not code or not name:
                continue
            level, parent_digits, digits = _infer(code)
            if level == 0:
                continue

            parent_code = parent_digits if parent_digits else None
            parent_name = ancestors_name.get(parent_digits) if parent_digits else None

            # build level_path: sheet|code at each level
            parts = [sheet_name]
            if level >= 2:
                parts.append(digits[:2])
            if level >= 3:
                parts.append(digits[:4])
            if level >= 4:
                parts.append(digits[:6])
            level_path = "|".join(parts)

            fin_class = _to_str(row[C_FIN]) if C_FIN < len(row) else ""
            unit = _to_str(row[C_UNIT]) if C_UNIT < len(row) else ""
            p3 = _norm_price(row[C_PRICE3]) if C_PRICE3 < len(row) else None
            p2 = _norm_price(row[C_PRICE2]) if C_PRICE2 < len(row) else None
            p1 = _norm_price(row[C_PRICE1]) if C_PRICE1 < len(row) else None
            content = _to_str(row[C_CONTENT]) if C_CONTENT < len(row) else ""
            exclude = _to_str(row[C_EXCLUDE]) if C_EXCLUDE < len(row) else ""
            remark = _to_str(row[C_REMARK]) if C_REMARK < len(row) else ""

            cur.execute(
                """INSERT OR REPLACE INTO sn_ms_codes
                   (code, p_code, name, level, sheet_name, sheet_title,
                    level_path, fin_class, unit, price_l1, price_l2, price_l3,
                    content, exclude, remark)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (code, parent_code, name, level, sheet_name, sheet_title,
                 level_path, fin_class or None, unit or None, p1, p2, p3,
                 content or None, exclude or None, remark or None),
            )
            inserted += 1
            # Cache for parent_name resolution
            if level <= 4:
                ancestors_name[code] = name

    conn.commit()
    wb.close()
    return inserted


def main(argv=None):
    p = argparse.ArgumentParser()
    p.add_argument("--xlsx", default=XLSX)
    p.add_argument("--db", default=None)
    p.add_argument("--truncate", action="store_true")
    args = p.parse_args(argv)

    db_path = args.db or DB_PATH

    if not os.path.exists(args.xlsx):
        print(f"xlsx not found: {args.xlsx}", file=sys.stderr)
        return 2
    if not os.path.exists(db_path):
        print(f"db not found: {db_path}", file=sys.stderr)
        return 2

    sql_path = os.path.join(os.path.dirname(os.path.abspath(db_path)), "sn_ms_schema.sql")
    conn = sqlite3.connect(db_path)
    if os.path.exists(sql_path):
        with open(sql_path, "r", encoding="utf-8") as fh:
            conn.executescript(fh.read())

    n = ingest(args.xlsx, conn, truncate=args.truncate)
    conn.close()
    print(f"inserted/updated {n} rows into sn_ms_codes")
    return 0


if __name__ == "__main__":
    sys.exit(main())
