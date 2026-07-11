"""Ingest 1-15 batch XLSX files into the unified database."""
from __future__ import annotations
import os
import re
import json
import argparse
from typing import Optional
from openpyxl import load_workbook

from . import db

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BATCH_DIRS = ["01-06批", "07-15批", "16批", "17-第十七批-药品限适应症(抗肿瘤+肌肉骨骼)", "18-第十八批-药品限适应症(神经系统药物)"]
SOURCE = "nhsa_batch"

# Per-batch public-date / announcement-url map (from batches_attachments.json + batches_1to6.json)
BATCH_META = {
    "第一批":   ("2025-05-23", "https://www.nhsa.gov.cn/art/2025/5/23/art_109_16625.html"),
    "第二批":   ("2025-07-22", "https://www.nhsa.gov.cn/art/2025/7/22/art_109_17342.html"),
    "第三批":   ("2025-07-30", "https://www.nhsa.gov.cn/art/2025/7/30/art_109_17438.html"),
    "第四批":   ("2025-08-08", "https://www.nhsa.gov.cn/art/2025/8/8/art_109_17527.html"),
    "第五批":   ("2025-08-14", "https://www.nhsa.gov.cn/art/2025/8/14/art_109_17567.html"),
    "第六批":   ("2025-08-26", "https://www.nhsa.gov.cn/art/2025/8/26/art_109_17691.html"),
    "第七批":   ("2025-12-12", "https://www.nhsa.gov.cn/art/2025/12/12/art_109_19013.html"),
    "第八批":   ("2026-01-06", "https://www.nhsa.gov.cn/art/2026/1/6/art_109_19250.html"),
    "更新发布": ("2026-01-28", "https://www.nhsa.gov.cn/art/2026/1/28/art_109_19496.html"),
    "第九批":   ("2026-04-21", "https://www.nhsa.gov.cn/art/2026/4/21/art_109_20277.html"),
    "第十批":   ("2026-04-28", "https://www.nhsa.gov.cn/art/2026/4/28/art_109_20351.html"),
    "第十一批": ("2026-05-11", "https://www.nhsa.gov.cn/art/2026/5/11/art_109_20462.html"),
    "第十二批": ("2026-05-23", "https://www.nhsa.gov.cn/art/2026/5/23/art_109_20682.html"),
    "第十三批": ("2026-06-01", "https://www.nhsa.gov.cn/art/2026/6/1/art_109_20822.html"),
    "第十四批": ("2026-06-09", "https://www.nhsa.gov.cn/art/2026/6/9/art_109_20899.html"),
    "第十五批": ("2026-06-16", "https://www.nhsa.gov.cn/art/2026/6/16/art_109_21004.html"),
    "第十六批": ("2026-06-22", "https://www.nhsa.gov.cn/art/2026/6/22/art_109_21056.html"),
    "第十七批": ("2026-06-29", "https://www.nhsa.gov.cn/art/2026/6/29/art_109_21146.html"),
    "第十八批": ("2026-07-10", "https://www.nhsa.gov.cn/art/2026/7/10/art_109_21366.html"),
}

CHINESE_QUOTES = "“”"


def extract_batch_label(folder_name: str) -> Optional[str]:
    """e.g. '07-第七批-医疗服务项目重复收费' -> '第七批' ; '2026-01-28更新发布-药品8项规则' -> '更新发布'"""
    m = re.search(r"(第[一二三四五六七八九十]+批|更新发布)", folder_name)
    return m.group(1) if m else None


def extract_rule_subject(file_label: str) -> str:
    """Pull the canonical rule name from the file label.
    Prefer the content inside Chinese curly quotes; fall back to prefix/suffix stripping."""
    m = re.search(r"[" + CHINESE_QUOTES + r"]([^" + CHINESE_QUOTES + r"]+)[" + CHINESE_QUOTES + r"]", file_label)
    if m:
        return m.group(1).strip()
    s = file_label
    s = re.sub(r"^第[一二三四五六七八九十]+批(?:[-－]\d+)?[.．、]?\s*", "", s)
    s = re.sub(r"\s*规则(?:主项目|子项目)?对应(?:部分)?知识点明细\.(?:pdf|xlsx)$", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\.(?:pdf|xlsx)$", "", s, flags=re.IGNORECASE)
    s = re.sub(r"^\d+[-－]\s*", "", s)
    return s.strip()


def infer_object_type(file_label: str, rule_subject: str) -> str:
    s = (file_label + " " + rule_subject).lower()
    if "中药饮片" in s:
        return "tcm"
    if "耗材" in s:
        return "consumable"
    if "重复收费" in s or "分解收费" in s or "项目a" in s and "项目b" in s:
        return "pair"
    if "药品" in s or rule_subject.endswith(("使用", "保险", "用药", "用药安全", "专用", "禁用")):
        return "drug"
    return "service"


def infer_category(rule_subject: str) -> Optional[str]:
    s = rule_subject
    if any(k in s for k in ("工伤", "生育", "二线", "适应症", "支付疗程", "医疗机构级别", "就医方式", "互联网医院", "折价收费", "重复收费", "分解收费", "限定频次", "限年龄", "周期超频次")):
        if "中药" in s or "药品" in s or any(k in s for k in ("工伤", "生育", "二线", "适应症", "支付疗程", "医疗机构级别", "就医方式", "互联网医院")):
            return "政策-药品" if any(k in s for k in ("工伤", "生育", "二线", "适应症", "支付疗程", "医疗机构级别", "就医方式", "互联网医院", "中药")) else None
        return "政策-服务"
    if any(k in s for k in ("儿童专用", "儿童禁用", "区分性别", "超说明书", "禁忌", "超量", "超大处方", "重复开药", "相互作用", "配伍", "老年人", "妊娠期")):
        return "医疗-药品"
    return None


# ---------- column-header fingerprint ----------

def col_role(header_text: str) -> Optional[str]:
    """Map a raw header cell (may include newlines) to a canonical role."""
    if not header_text:
        return None
    h = re.sub(r"\s+", "", header_text)
    if h in ("序号", "对应知识点序号", "知识点序号"):
        return "seq"
    if h in ("药品通用名",):
        return "subject_name"
    if h in ("医疗服务项目名称",):
        return "subject_name"
    if h.startswith("医疗服务项目A名称") or h.startswith("中药饮片A名称"):
        return "subject_name"
    if h.startswith("医疗服务项目B名称") or h.startswith("中药饮片B名称"):
        return "subject_name_b"
    if h in ("中药饮片名称",):
        return "subject_name"
    if h.startswith("项目A名称") or h.startswith("项目B名称"):
        return "subject_name"
    if h in ("检出逻辑",):
        return "detection_logic"
    if h.startswith("逻辑依据"):
        return "logic_basis"
    if "知识点对应" in h and "数量" in h:
        return "code_count"
    if h.startswith("限定性别") or h == "限定性别":
        return "remark"
    if h in ("药品代码",) or h.startswith("项目代码") or h.startswith("药品代码") or h.startswith("中药饮片代码") or h.startswith("医疗服务项目代码"):
        return "codes"
    if h.startswith("医疗服务项目A代码") or h.startswith("中药饮片A代码"):
        return "codes"
    if h.startswith("医疗服务项目B代码") or h.startswith("中药饮片B代码"):
        return "codes_b"
    if h in ("时间区间", "单日上限数量"):
        return "remark"
    if h in ("备注",):
        return "remark"
    return None


def read_xlsx_rows(path: str):
    """Read xlsx and return (header_map, data_rows, codes_by_seq).

    Always reads the FIRST sheet (the KP-definition sheet) for data_rows.
    Earlier versions used wb.active which is unreliable: if the last-active
    sheet is the codes sheet, every column ends up NULL.
    If a second sheet contains code-like columns, group its codes by seq.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = wb.sheetnames
    kp_sheet = None
    for sn in sheets:
        head = ""
        ws_p = wb[sn]
        for r in ws_p.iter_rows(max_row=6, values_only=True):
            for c in r:
                if c is not None:
                    head += str(c)
        if "序号" in head and (
            "药品通用名" in head or "医疗服务项目" in head
            or "中药饮片名称" in head or "药品名称" in head
            or "项目名称" in head
        ):
            kp_sheet = sn
            break
    if kp_sheet is None:
        kp_sheet = sheets[0]
    header_map, data = _parse_sheet(wb[kp_sheet])
    codes_by_seq = {}
    for sn in sheets:
        if sn == kp_sheet:
            continue
        head2 = ""
        for r in wb[sn].iter_rows(max_row=6, values_only=True):
            for c in r:
                if c is not None:
                    head2 += str(c)
        if "药品代码" in head2 or "项目代码" in head2 or "中药饮片代码" in head2:
            _parse_codes_sheet(wb[sn], codes_by_seq)
            break
    wb.close()
    return header_map, data, codes_by_seq


def _parse_sheet(ws):
    rows = list(ws.iter_rows(values_only=False))
    header_idx = None
    for i, row in enumerate(rows[:8], start=1):
        text = "".join(str(c.value) for c in row if c.value is not None)
        if "序号" in text and (
            "检出逻辑" in text or "药品通用名" in text
            or "医疗服务项目" in text or "中药饮片名称" in text
            or "药品名称" in text or "项目名称" in text
        ):
            header_idx = i
            break
    if header_idx is None:
        return None, []
    header_cells = rows[header_idx - 1]
    header_map = {}
    for c in header_cells:
        if c.value is None:
            continue
        role = col_role(str(c.value))
        if role:
            header_map[c.column_letter] = role
    data = []
    for row in rows[header_idx:]:
        rec = {"_cells": {}}
        for c in row:
            v = c.value
            if v is None:
                continue
            role = header_map.get(c.column_letter)
            if role is None:
                continue
            rec["_cells"][role] = v
        if rec["_cells"]:
            data.append(rec)
    return header_map, data


def _parse_codes_sheet(ws, out: dict):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return
    header_idx = None
    for i, row in enumerate(rows[:8], start=1):
        text = "".join(str(c) for c in row if c is not None)
        if "序号" in text and ("药品代码" in text or "项目代码" in text or "中药饮片代码" in text):
            header_idx = i
            break
    if header_idx is None:
        return
    hdr = list(rows[header_idx - 1])

    def find_col(*needles):
        for j, h in enumerate(hdr):
            if h is None:
                continue
            s = re.sub(r"\s+", "", str(h))
            for n in needles:
                if n in s:
                    return j
        return None

    kp_seq_col = find_col("对应知识点序号", "知识点序号")
    code_col = find_col("药品代码", "项目代码", "中药饮片代码")
    if code_col is None:
        return
    cur = None
    for row in rows[header_idx:]:
        if kp_seq_col is not None and kp_seq_col < len(row) and row[kp_seq_col] is not None:
            try:
                cur = int(str(row[kp_seq_col]).strip())
            except (ValueError, TypeError):
                cur = None
        if cur is None:
            continue
        code = row[code_col] if code_col < len(row) else None
        if code is None:
            continue
        code = str(code).strip()
        if not code:
            continue
        out.setdefault(cur, []).append(code)



def to_int(v) -> Optional[int]:
    if v is None:
        return None
    try:
        return int(str(v).strip())
    except (ValueError, TypeError):
        try:
            return int(float(v))
        except (ValueError, TypeError):
            return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--reset", action="store_true", help="Reset DB before ingest")
    args = ap.parse_args()

    if args.reset:
        db.reset_db()
    else:
        db.init_db()

    total_kp = 0
    files_done = 0
    with db.connect() as conn:
        for sub in BATCH_DIRS:
            base = os.path.join(PROJECT_ROOT, sub)
            if not os.path.isdir(base):
                continue
            for entry in sorted(os.listdir(base)):
                full = os.path.join(base, entry)
                if not os.path.isdir(full):
                    continue
                batch_label = extract_batch_label(entry)
                if not batch_label:
                    print(f"  skip: cannot extract batch label from {entry}")
                    continue
                pub_date, ann_url = BATCH_META.get(batch_label, (None, None))
                for fname in sorted(os.listdir(full)):
                    if not fname.lower().endswith(".xlsx"):
                        continue
                    xlsx_path = os.path.join(full, fname)
                    pdf_path = os.path.splitext(xlsx_path)[0] + ".pdf"
                    pdf_path = pdf_path if os.path.exists(pdf_path) else None
                    subject = extract_rule_subject(fname)
                    object_type = infer_object_type(fname, subject)
                    category = infer_category(subject)

                    batch_id = db.get_or_create_batch(
                        conn, SOURCE, batch_label,
                        pub_date=pub_date, ann_url=ann_url,
                        pdf_path=pdf_path, xlsx_path=xlsx_path,
                    )
                    rule_id = db.get_or_create_rule(
                        conn, SOURCE, subject, batch_id,
                        category=category, object_type=object_type, xlsx_path=xlsx_path,
                    )
                    db.replace_kp_for_rule(conn, rule_id)

                    header_map, rows, codes_by_seq = read_xlsx_rows(xlsx_path)
                    if not rows:
                        print(f"  no data: {fname}")
                        continue
                    n = 0
                    for r in rows:
                        cells = r["_cells"]
                        # skip 合计/总计 summary rows. 25 affected files place 合计 in the 序号 column
                        # (not in subject_name). Detection logic frequently contains 总计 legitimately
                        # (e.g. "数量总计超过上限"), so check only seq + subject_name, not all cells.
                        subj = cells.get("subject_name")
                        seq_cell = cells.get("seq")
                        if (subj and "合计" in str(subj)) or (isinstance(seq_cell, str) and "合计" in seq_cell):
                            continue
                        seq = to_int(cells.get("seq"))
                        subject_name = db.normalize_text(subj)
                        code_count = to_int(cells.get("code_count"))
                        detection_logic = cells.get("detection_logic")
                        logic_basis = cells.get("logic_basis")
                        codes = cells.get("codes")
                        remark = cells.get("remark")
                        raw = json.dumps(cells, ensure_ascii=False, default=str)
                        # Drop narrative rows: e.g. "说明：药品区分性别使用...".
                        # The first non-empty cell begins with 说明/通知/公告.
                        joined_first = (str(subject_name or "") + str(seq_cell or "")).lstrip()
                        if joined_first.startswith(("说明", "通知", "公告", "备注")):
                            continue
                        # pair rules (重复收费 etc.): subject_name is "A vs B" and codes is A+B
                        subj_b = db.normalize_text(cells.get("subject_name_b"))
                        if subj_b:
                            s = db.normalize_text(subject_name) or ""
                            subject_name = f"{s} vs {subj_b}"
                        else:
                            subject_name = db.normalize_text(subject_name)
                        raw_codes = []
                        cn = db.normalize_text(codes)
                        if cn:
                            raw_codes.append(cn)
                        cb = db.normalize_text(cells.get("codes_b"))
                        if cb:
                            raw_codes.append(cb)
                        kp_id = db.insert_kp(
                            conn, rule_id,
                            seq=seq,
                            subject_name=db.normalize_text(subject_name),
                            code_count=code_count,
                            detection_logic=db.normalize_text(detection_logic),
                            logic_basis=db.normalize_text(logic_basis),
                            codes=None,  # codes now live in knowledge_point_codes (1:N)
                            remark=db.normalize_text(remark),
                            raw_row=db.normalize_text(raw),
                        )
                        # link codes: prefer codes sheet (multi-code drugs); fall back to inline A/B codes
                        if seq is not None and seq in codes_by_seq:
                            db.insert_kp_codes(conn, kp_id, codes_by_seq[seq])
                        elif raw_codes:
                            db.insert_kp_codes(conn, kp_id, raw_codes)
                        # mirror back to knowledge_points.codes (joined with 、 for FTS) for FTS / legacy
                        conn.execute("UPDATE knowledge_points SET codes = (SELECT GROUP_CONCAT(code, char(12539)) FROM knowledge_point_codes WHERE kp_id = ?) WHERE id = ?", (kp_id, kp_id))
                        n += 1
                    db.update_rule_row_count(conn, rule_id)
                    total_kp += n
                    files_done += 1
                    print(f"  {batch_label}  {subject}  ->  {n} KPs  ({fname[:40]})")

    print(f"\n=== XLSX ingest done: {files_done} files, {total_kp} knowledge points ===")
    print(f"DB: {db.DB_PATH}")


if __name__ == "__main__":
    main()
