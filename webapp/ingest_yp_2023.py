"""
2023 版药品目录入库脚本
解析 3 个 XLSX 入库到 webapp/data/kp.db:
- yp_catalog_2023       (3,732 目录条目,5 类)
- yp_catalog_adjustments (507 调整记录)
- yp_catalog_listing    (344 挂网信息,含 15 位 yp_code)

幂等,可重复执行(--reset 清空重建)。
"""
import argparse
import sqlite3
import re
from pathlib import Path

import openpyxl

DB_PATH = Path(__file__).parent / "data" / "kp.db"
DEFAULT_DIR = Path(__file__).parent.parent / "2023版药品目录"
VERSION = "2023"

# ---- SQL 建表 ----
DDL = [
    """CREATE TABLE IF NOT EXISTS yp_catalog_2023 (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        code TEXT,
        name TEXT NOT NULL,
        category TEXT NOT NULL,
        tcm_payment INTEGER,
        category_code TEXT,
        subcategory TEXT,
        dosage_form TEXT,
        list_class TEXT,
        payment_standard TEXT,
        payment_validity TEXT,
        remark TEXT,
        spec_count INTEGER DEFAULT 0,
        manufacturer_count INTEGER DEFAULT 0,
        version TEXT,
        sheet_source TEXT,
        UNIQUE(name, category, version)
    )""",
    """CREATE INDEX IF NOT EXISTS idx_yp_cat_category ON yp_catalog_2023(category)""",
    """CREATE INDEX IF NOT EXISTS idx_yp_cat_list_class ON yp_catalog_2023(list_class)""",
    """CREATE INDEX IF NOT EXISTS idx_yp_cat_code ON yp_catalog_2023(code)""",
    """CREATE TABLE IF NOT EXISTS yp_catalog_adjustments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        change_type TEXT NOT NULL,
        drug_name TEXT NOT NULL,
        drug_category TEXT,
        change_detail TEXT,
        list_no TEXT,
        version TEXT,
        source_sheet TEXT,
        UNIQUE(change_type, drug_name, version)
    )""",
    """CREATE TABLE IF NOT EXISTS yp_catalog_listing (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        source_list TEXT NOT NULL,
        access_type TEXT,
        generic_seq TEXT,
        drug_name TEXT NOT NULL,
        dosage_form TEXT,
        spec TEXT,
        yp_code TEXT,
        min_pkg_qty TEXT,
        min_prep_unit TEXT,
        min_pkg_unit TEXT,
        packaging TEXT,
        manufacturer TEXT,
        version TEXT,
        UNIQUE(yp_code, source_list, version)
    )""",
    """CREATE INDEX IF NOT EXISTS idx_yp_list_ypcode ON yp_catalog_listing(yp_code)""",
    """CREATE INDEX IF NOT EXISTS idx_yp_list_name ON yp_catalog_listing(drug_name)""",
]

FTS_DDL = [
    """CREATE VIRTUAL TABLE IF NOT EXISTS yp_catalog_fts USING fts5(
        name, subcategory, category_code, dosage_form, remark,
        content="yp_catalog_2023", content_rowid="id"
    )""",
    # 同步触发器
    """CREATE TRIGGER IF NOT EXISTS yp_cat_ai AFTER INSERT ON yp_catalog_2023 BEGIN
        INSERT INTO yp_catalog_fts(rowid, name, subcategory, category_code, dosage_form, remark)
        VALUES (new.id, new.name, IFNULL(new.subcategory,""), IFNULL(new.category_code,""),
                IFNULL(new.dosage_form,""), IFNULL(new.remark,""));
    END""",
    """CREATE TRIGGER IF NOT EXISTS yp_cat_ad AFTER DELETE ON yp_catalog_2023 BEGIN
        INSERT INTO yp_catalog_fts(yp_catalog_fts, rowid, name, subcategory, category_code, dosage_form, remark)
        VALUES ("delete", old.id, old.name, IFNULL(old.subcategory,""), IFNULL(old.category_code,""),
                IFNULL(old.dosage_form,""), IFNULL(old.remark,""));
    END""",
    """CREATE TRIGGER IF NOT EXISTS yp_cat_au AFTER UPDATE ON yp_catalog_2023 BEGIN
        INSERT INTO yp_catalog_fts(yp_catalog_fts, rowid, name, subcategory, category_code, dosage_form, remark)
        VALUES ("delete", old.id, old.name, IFNULL(old.subcategory,""), IFNULL(old.category_code,""),
                IFNULL(old.dosage_form,""), IFNULL(old.remark,""));
        INSERT INTO yp_catalog_fts(rowid, name, subcategory, category_code, dosage_form, remark)
        VALUES (new.id, new.name, IFNULL(new.subcategory,""), IFNULL(new.category_code,""),
                IFNULL(new.dosage_form,""), IFNULL(new.remark,""));
    END""",
]


def open_db():
    conn = sqlite3.connect(str(DB_PATH))
    return conn


def ensure_schema(conn, reset=False):
    cur = conn.cursor()
    for sql in DDL + FTS_DDL:
        cur.execute(sql)
    if reset:
        for tbl in ["yp_catalog_2023", "yp_catalog_adjustments", "yp_catalog_listing"]:
            cur.execute(f"DELETE FROM {tbl}")
            # 重建自增
            cur.execute(f"DELETE FROM sqlite_sequence WHERE name=?", (tbl,))
    conn.commit()


def cell(v):
    if v is None: return None
    s = str(v).strip()
    return s if s else None


# 各 sheet 的列位置(硬编码,header 不含 甲乙类 列名)
LAYOUTS = {
    "西药":       {"class": 5, "code": 6, "name": 7, "form": 8, "remark": 9, "payment": None, "validity": None},
    "中成药":     {"class": 5, "code": 6, "name": 7, "form": None, "remark": 8, "payment": None, "validity": None},
    "谈判西药":   {"class": 5, "code": 6, "name": 7, "form": None, "remark": 10, "payment": 8, "validity": 11},
    "谈判中成药": {"class": 5, "code": 6, "name": 7, "form": None, "remark": 10, "payment": 8, "validity": 11},
    "竞价药品":   {"class": 5, "code": 6, "name": 7, "form": None, "remark": 9, "payment": 8, "validity": 10},
}


def parse_main_sheet(conn, xlsx_path, sheet_name, category):
    """按 LAYOUTS 中的硬编码列位置解析。"""
    layout = LAYOUTS.get(category)
    if not layout:
        print(f"  WARN: 未知 category {category}")
        return 0
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=4, values_only=True))
    wb.close()
    cur = conn.cursor()
    n = 0
    inserted = 0
    for r in rows:
        if not r:
            continue
        # 关键列:必须有 名称
        if layout["name"] >= len(r):
            continue
        name = cell(r[layout["name"]])
        if not name:
            continue
        # code 不是必须(中药饮片按 序号)
        code = cell(r[layout["code"]]) if layout["code"] < len(r) else None
        if not code:
            continue
        list_class = cell(r[layout["class"]]) if layout["class"] < len(r) else None
        dosage_form = cell(r[layout["form"]]) if layout["form"] is not None and layout["form"] < len(r) else None
        remark = cell(r[layout["remark"]]) if layout["remark"] is not None and layout["remark"] < len(r) else None
        payment_standard = cell(r[layout["payment"]]) if layout["payment"] is not None and layout["payment"] < len(r) else None
        payment_validity = cell(r[layout["validity"]]) if layout["validity"] is not None and layout["validity"] < len(r) else None
        cur.execute("""INSERT OR IGNORE INTO yp_catalog_2023
            (code, name, category, list_class, category_code, subcategory, dosage_form,
             payment_standard, payment_validity, remark, version, sheet_source)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (code, name, category, list_class, None, None, dosage_form,
             payment_standard, payment_validity, remark, VERSION, sheet_name))
        n += 1
        if cur.rowcount > 0:
            inserted += 1
    conn.commit()
    print(f"  {sheet_name} -> category={category}: scanned={n}, inserted={inserted}")
    return inserted


def parse_tcm_paid(conn, xlsx_path):
    """中药饮片-支付 sheet:2 列布局(col 0-2 + col 3-5)"""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["中药饮片部分-支付892"]
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    wb.close()
    cur = conn.cursor()
    inserted = 0
    for r in rows:
        if not r: continue
        # Left half
        for base in (0, 3):
            code = cell(r[base])
            name = cell(r[base+1])
            remark = cell(r[base+2])
            if not (code and name):
                continue
            cur.execute("""INSERT OR IGNORE INTO yp_catalog_2023
                (code, name, category, tcm_payment, list_class, remark, version, sheet_source)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                (code, name, "中药饮片", 1, None, remark, VERSION, "中药饮片部分-支付892"))
            if cur.rowcount > 0:
                inserted += 1
    conn.commit()
    print(f"  中药饮片-支付: inserted={inserted}")
    return inserted


def parse_tcm_unpaid(conn, xlsx_path):
    """中药饮片-不支付 sheet:第 2 行 col 0 是长文本"""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["中药饮片部分-不支付"]
    rows = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    wb.close()
    if not rows or not rows[0] or not rows[0][0]:
        return 0
    text = str(rows[0][0])
    # 移除空白和换行符,按 、 和 , 分割
    cleaned = text.replace("\n", " ").replace("\r", " ")
    items = re.split(r"[、,,;;  ]+", cleaned)
    items = [x.strip(" ()（）") for x in items if x.strip(" ()（）")]
    cur = conn.cursor()
    inserted = 0
    for i, name in enumerate(items, start=1):
        cur.execute("""INSERT OR IGNORE INTO yp_catalog_2023
            (code, name, category, tcm_payment, version, sheet_source)
            VALUES (?, ?, ?, ?, ?, ?)""",
            (str(i), name, "中药饮片", 0, VERSION, "中药饮片部分-不支付"))
        if cur.rowcount > 0:
            inserted += 1
    conn.commit()
    print(f"  中药饮片-不支付: items={len(items)}, inserted={inserted}")
    return inserted


# XLSX 3 各 sheet 的列布局(基于实际 header 第 3 行)
# 每行:(seq_col, cat_col, name_col, detail_col, list_no_col)
ADJ_LAYOUTS = {
    "1.直接调出-1": ("直接调出", 0, 1, 2, None, None),
    "2.谈判转常规目录-51": ("谈判转常规", 0, 1, 2, 4, 5),
    "3.新增纳入-126": ("新增纳入", 0, 1, 3, 5, None),
    "4.常规目录调整支付范围-225（223+2）": ("调整支付范围", 0, 1, 3, 6, 8),
    "5.协议期内谈判药修订支付范围-102": ("协议期内谈判药修订支付范围", 0, None, 1, 4, None),
    "6.规范名称-2": ("规范名称", 0, 1, 2, 4, None),
    # 过程稿不收录
}


def parse_adjustments(conn, xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    cur = conn.cursor()
    inserted = 0
    for sn in wb.sheetnames:
        layout = ADJ_LAYOUTS.get(sn)
        if not layout:
            continue
        change_type, seq_col, cat_col, name_col, detail_col, list_no_col = layout
        ws = wb[sn]
        rows = list(ws.iter_rows(min_row=4, values_only=True))
        for r in rows:
            if not r or seq_col is None or seq_col >= len(r):
                continue
            seq = cell(r[seq_col])
            if not seq:
                continue
            cat = cell(r[cat_col]) if cat_col is not None and cat_col < len(r) else None
            name = cell(r[name_col]) if name_col is not None and name_col < len(r) else None
            if not name:
                continue
            detail = cell(r[detail_col]) if detail_col is not None and detail_col < len(r) else None
            list_no = cell(r[list_no_col]) if list_no_col is not None and list_no_col < len(r) else None
            cur.execute("""INSERT OR IGNORE INTO yp_catalog_adjustments
                (change_type, drug_name, drug_category, change_detail, list_no, version, source_sheet)
                VALUES (?, ?, ?, ?, ?, ?, ?)""",
                (change_type, name, cat, detail, list_no, VERSION, sn))
            if cur.rowcount > 0:
                inserted += 1
    wb.close()
    conn.commit()
    print(f"  yp_catalog_adjustments: inserted={inserted}")
    return inserted


def parse_listing(conn, xlsx_path):
    """XLSX 4:谈判/竞价/集采 3 个挂网 sheet"""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    cur = conn.cursor()
    inserted = 0
    sheet_to_source = {
        "药品目录挂网信息表-谈判药品": "谈判",
        "药品目录挂网信息表-竞价药品": "竞价",
        "药品目录挂网信息表-集采药品": "集采",
    }
    for sn, source_list in sheet_to_source.items():
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        rows = list(ws.iter_rows(min_row=4, values_only=True))
        for r in rows:
            if not r: continue
            seq = cell(r[0])
            generic_seq = cell(r[1])
            access_type = cell(r[2])
            name = cell(r[3])
            dosage_form = cell(r[4])
            spec = cell(r[5])
            yp_code = cell(r[6])
            min_pkg_qty = cell(r[7])
            min_prep_unit = cell(r[8])
            min_pkg_unit = cell(r[9])
            packaging = cell(r[10])
            manufacturer = cell(r[11])
            if not name: continue
            cur.execute("""INSERT OR IGNORE INTO yp_catalog_listing
                (source_list, access_type, generic_seq, drug_name, dosage_form, spec, yp_code,
                 min_pkg_qty, min_prep_unit, min_pkg_unit, packaging, manufacturer, version)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (source_list, access_type, generic_seq, name, dosage_form, spec, yp_code,
                 min_pkg_qty, min_prep_unit, min_pkg_unit, packaging, manufacturer, VERSION))
            if cur.rowcount > 0:
                inserted += 1
    wb.close()
    conn.commit()
    print(f"  yp_catalog_listing: inserted={inserted}")
    return inserted


def compute_spec_counts(conn):
    """对每个 yp_catalog_2023 行,统计 yp_codes 里 name 模糊匹配的厂家规格/厂家数。
    性能策略:Python 端预加载 yp_codes.reg_name + manufacturer 字典,
    再批量 UPDATE(每 500 条提交一次)。"""
    cur = conn.cursor()
    print("  loading yp_codes.reg_name/manufacturer ...", flush=True)
    rows = cur.execute("SELECT reg_name, manufacturer FROM yp_codes WHERE reg_name IS NOT NULL").fetchall()
    print(f"  loaded {len(rows)} yp_codes rows", flush=True)
    # 预处理:nospace(reg_name) -> set(manufacturer)
    from collections import defaultdict
    bucket = defaultdict(lambda: {"specs": 0, "mans": set()})
    for reg, man in rows:
        key = (reg or "").replace(" ", "").replace("\u3000", "")
        if not key: continue
        bucket[key]["specs"] += 1
        if man: bucket[key]["mans"].add(man.strip())
    # 取目录
    cats = cur.execute("SELECT id, name FROM yp_catalog_2023").fetchall()
    print(f"  matching {len(cats)} catalog rows ...", flush=True)
    updated = 0
    BATCH = 500
    batch = []
    for cid, cname in cats:
        key = (cname or "").replace(" ", "").replace("\u3000", "")
        # 找最长前缀匹配(防止「人」误匹配所有「人」字头药)
        best = None
        best_len = 0
        for k in bucket.keys():
            if len(k) > best_len and key and (key in k or k in key):
                if len(k) >= 2:  # 至少 2 字避免噪声
                    best = k; best_len = len(k)
        if best:
            stats = bucket[best]
            batch.append((stats["specs"], len(stats["mans"]), cid))
        else:
            batch.append((0, 0, cid))
        if len(batch) >= BATCH:
            cur.executemany("UPDATE yp_catalog_2023 SET spec_count=?, manufacturer_count=? WHERE id=?", batch)
            conn.commit()
            updated += len(batch)
            batch = []
    if batch:
        cur.executemany("UPDATE yp_catalog_2023 SET spec_count=?, manufacturer_count=? WHERE id=?", batch)
        conn.commit()
        updated += len(batch)
    print(f"  updated {updated} rows", flush=True)
    n = cur.execute("SELECT COUNT(*) FROM yp_catalog_2023 WHERE spec_count > 0").fetchone()[0]
    n_zero = cur.execute("SELECT COUNT(*) FROM yp_catalog_2023 WHERE spec_count=0").fetchone()[0]
    print(f"  spec_count > 0: {n}, = 0: {n_zero}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dir", default=str(DEFAULT_DIR))
    ap.add_argument("--reset", action="store_true")
    args = ap.parse_args()
    base = Path(args.dir)
    print(f"DB: {DB_PATH}")
    print(f"DIR: {base}")
    conn = open_db()
    ensure_schema(conn, reset=args.reset)
    print()
    print("[1/4] 入库 XLSX 2 主目录 (5 大类)")
    n = 0
    n += parse_main_sheet(conn, base / "2.2023年版药品目录（含保密价格）.xlsx", "西药部分1333", "西药")
    n += parse_main_sheet(conn, base / "2.2023年版药品目录（含保密价格）.xlsx", "中成药部分1323", "中成药")
    n += parse_main_sheet(conn, base / "2.2023年版药品目录（含保密价格）.xlsx", "协议期内谈判药品部分-西药332", "谈判西药")
    n += parse_main_sheet(conn, base / "2.2023年版药品目录（含保密价格）.xlsx", "协议期内谈判药品部分-中成药67", "谈判中成药")
    n += parse_tcm_paid(conn, base / "2.2023年版药品目录（含保密价格）.xlsx")
    n += parse_tcm_unpaid(conn, base / "2.2023年版药品目录（含保密价格）.xlsx")
    print(f"  yp_catalog_2023 inserted: {n}")
    print()
    print("[2/4] 入库 XLSX 3 调整记录")
    parse_adjustments(conn, base / "3.2023年国家医保目录调整变更药品名单（谈判转常规改为51种）.xlsx")
    print()
    print("[3/4] 入库 XLSX 4 挂网信息")
    parse_listing(conn, base / "4.2023年医保目录新增药品挂网信息表（谈判、竞价、集采）.xlsx")
    print()
    print("[4/4] 计算 spec_count / manufacturer_count")
    compute_spec_counts(conn)
    print()
    print("=== 入库完成 ===")
    for t in ("yp_catalog_2023", "yp_catalog_adjustments", "yp_catalog_listing", "yp_catalog_fts"):
        try:
            n = conn.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
            print(f"  {t}: {n:>8,} 行")
        except Exception as e:
            print(f"  {t}: ERR {e}")
    conn.close()


if __name__ == "__main__":
    main()
